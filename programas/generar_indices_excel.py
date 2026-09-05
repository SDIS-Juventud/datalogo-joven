# -*- coding: utf-8 -*-
"""
Reconstruye los dos Indice.xlsx a partir de data/ecosistema-documentos.json.

Por que existe este paso: hasta septiembre de 2026 habia dos listas del mismo
repositorio, el JSON que alimenta la pagina y los Excel de cada carpeta, y nadie
sincronizaba la segunda. El indice de Ecosistema general llego a tener 30 filas
cuando la pagina ya mostraba 34 documentos. Ahora el JSON es la unica fuente y
el Excel se genera desde ahi, asi que no se puede volver a desfasar.

Los Excel NO se editan a mano: lo que se edite se pierde en la siguiente corrida.
Para cambiar algo, se cambia el JSON y se vuelve a correr.

    python programas/generar_indices_excel.py
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

raiz = Path(__file__).resolve().parent.parent

# (encabezado, campo del JSON, ancho de la columna)
COLUMNAS_COMUNES = [
    ("No.", "number", 6),
    ("Título", "title", 58),
    ("Autor", "author", 34),
    ("Año de publicación", "year", 14),
]
COLUMNA_CATEGORIA = ("Categoría de enfoque diferencial", "category", 26)
COLUMNAS_FINALES = [
    ("Ideas principales", "ideas", 90),
    ("Archivo PDF", None, 52),
]


def _escribir(carpeta, fichas):
    columnas = list(COLUMNAS_COMUNES)
    if carpeta == "Enfoque diferencial":
        columnas.append(COLUMNA_CATEGORIA)
    columnas += COLUMNAS_FINALES

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    for i, (encabezado, _, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=i, value=encabezado)
        celda.font = Font(bold=True)
        celda.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = ancho

    for fila, ficha in enumerate(fichas, start=2):
        for i, (_, campo, _) in enumerate(columnas, start=1):
            # la ultima columna no sale de un campo: es el nombre del archivo
            valor = (Path(ficha["pdfPath"]).name if campo is None
                     else ficha.get(campo, ""))
            celda = ws.cell(row=fila, column=i, value=valor)
            celda.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"                      # el encabezado queda fijo al bajar
    ws.auto_filter.ref = ws.dimensions
    ruta = raiz / carpeta / "Indice.xlsx"
    wb.save(ruta)
    return ruta, len(fichas)


def generar():
    """Reescribe los dos indices y devuelve un resumen por carpeta."""
    documentos = json.loads(
        (raiz / "data" / "ecosistema-documentos.json").read_text(encoding="utf-8"))
    resumen = []
    for carpeta in ("Ecosistema general", "Enfoque diferencial"):
        fichas = [d for d in documentos if d["source"] == carpeta]
        _escribir(carpeta, fichas)
        sin_numero = sum(1 for d in fichas if not str(d.get("number", "")).strip())
        aviso = f"  ({sin_numero} sin consecutivo)" if sin_numero else ""
        resumen.append(f"{carpeta}/Indice.xlsx: {len(fichas)} documentos{aviso}")
    return resumen


if __name__ == "__main__":
    for linea in generar():
        print(linea)
